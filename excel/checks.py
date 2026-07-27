from openpyxl import load_workbook
from datetime import datetime

from excel.writer import get_sheet_name


def is_duplicate(workbook_path: str, service) -> bool:
    """
    Checks if a similar service already exists in the correct sheet.
    """

    workbook = load_workbook(workbook_path)

    sheet_name = get_sheet_name(service.date)

    sheet = workbook[sheet_name]

    for row in sheet.iter_rows(
        min_row=2,
        min_col=1,
        max_col=7, 
        values_only=True
    ):
        
        _, _, tag, service_name, address, phone, client = row

        if (
            str(service_name).strip() == str(service.service).strip()
            and str(tag).strip() == str(service.tag).strip()
            and str(client).strip() == str(service.client).strip()
            and str(phone).strip() == str(service.phone).strip()
            and str(address).strip() == str(service.address).strip()
        ):
            print(f"Duplicate found: {service_name}, {address}, {phone}, {client}")  # Debugging line
            return True

    return False